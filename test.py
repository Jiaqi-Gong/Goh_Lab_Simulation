import random
from datetime import datetime
from functools import partial
from typing import Tuple, List

import numpy as np  # numpy is required to make matrices
from PIL import Image
from numpy import ndarray
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time

# from SimulatorFile.Dynamic import DynamicSimulator
# from SimulatorFile.EnergyScan import EnergySimulator
from ExternalIO import *
import multiprocessing as mp
import os


def test_diamond():
    a = np.zeros((15, 15))
    length = 3
    n = length
    c = 0  # for count
    start = (1, 6)
    a[start[0]][start[1]] = -1

    pos = 1

    # make upper diamond
    for i in range(0, n + 1):
        print(i)
        for j in range(-c + 1, c):
            print(i, j, (start[0] + i, start[1] + j - 1))
            a[start[0] + i][start[1] + j - 1] = pos
            pos += 1

        c += 1
    # make lower diamond
    for i in range(n + 1, 2 * (n + 1) + 1):
        for j in range(-c + 1, c):
            a[start[0] + i][start[1] - j - 1] = pos
            pos += 1

        c -= 1

    print(a)


def test_random_choice():
    a = np.zeros((20, 20))

    surfaceWidth = 10
    surfaceLength = 10

    domainWidth = 3
    domainLength = 3

    x = random.randint(domainWidth, surfaceWidth - domainWidth)
    y = random.randint(domainLength, surfaceLength - domainLength)


class A():
    def __init__(self):
        self.aa = 1
        self.bb = 2
        self.cc = None
        for parameter in self.__dict__:
            if self.__dict__[parameter] == None:
                raise RuntimeError("parameter {} is not set".format(parameter))

    def check(self):
        print("self.aa" in locals())
        print("cc" in locals())

    def b(self):
        self.cc = 3


def p():
    p = 1


def t():
    print(p)


def test_excel():
    from openpyxl import Workbook
    import time

    book = Workbook()
    sheet = book.active

    sheet['A1'] = 56
    sheet['A2'] = 43

    now = time.strftime("%x")
    sheet['A3'] = now

    # book.save("sample.xlsx")
    name = "Simulation type {} trail {}_{}.xlsx".format(str(3), 1,
                                                        datetime.now().strftime("%H-%M-%S"))
    file_path = "Result/" + name
    book.save(file_path)


def _output():
    """
        Output the simulation result into a file
        Copy from old code with minor change
        """
    startTime = datetime.now()
    time.sleep(3)

    # calculate the time use
    time_consume = (datetime.now() - startTime)
    time_consume = time_consume.seconds

    # creates excel file
    wb = Workbook()
    ws1 = wb.create_sheet("Results", 0)

    # naming the columns in the worksheet
    ws1.cell(1, 1, "Surface Characteristics:")
    ws1.cell(1, 2, "BacteriaFile Characteristics:")
    ws1.cell(1, 3, "FilmFile Seed # ")
    ws1.cell(1, 4, "BacteriaFile Seed # ")
    ws1.cell(1, 5, "Min Energy:")
    ws1.cell(1, 6, "Min X:")
    ws1.cell(1, 7, "Min Y:")
    ws1.cell(1, 8, "Surface Charge at Min Energy:")
    ws1.cell(1, 9, "Min Energy Gradient Strip: ")
    ws1.cell(1, 10, "Histogram: ")

    # create numbering for histogram plot
    count = 0
    for i in range(10, 31):
        ws1.cell(1, i, count)
        ws1.cell(2, i, 0)
        count += 1

    # adjust column width to text length
    for i in range(ws1.max_column):
        text = ws1.cell(1, i + 1).value
        if type(text) == int:
            text = str(text)
            column_width = len(text) + 1
        else:
            column_width = len(text)
        ws1.column_dimensions[get_column_letter(i + 1)].width = column_width

    # split the result
    min_energy, min_x, min_y, min_energy_charge, min_charge, min_charge_x, min_charge_y = (91, 92, 93, 94, 95, 96, 97)

    # shows the gradient lines position
    grad_strip = min_x // 500

    # write the result
    ws1.cell(3, 1, str(1) + " : " + str(2))
    ws1.cell(3, 2, str(1) + " : " + str(2))
    ws1.cell(3, 3, 3)
    ws1.cell(3, 4, 4)
    ws1.cell(3, 5, min_energy)
    ws1.cell(3, 6, min_x)
    ws1.cell(3, 7, min_x)
    ws1.cell(3, 8, min_energy_charge)
    ws1.cell(3, 9, grad_strip)
    ws1.cell(4, 9, time_consume)

    # special count for simulation type 2
    # save the excel file into folder result
    name = "Simulation type {} trail {}_{}.xlsx".format(str(2), 1, datetime.now().strftime("%H-%M-%S"))
    file_path = "Result/" + name
    wb.save(file_path)


def test3D():
    length = 5
    width = 5
    height = 5
    # finds center of array
    center = int(np.floor(length / 2)), int(np.floor(width / 2)), int(np.floor(height / 2))
    radius = min(np.floor(length / 2), np.floor(width / 2), np.floor(height / 2))
    # indexes the array
    index_x, index_y, index_z = np.indices((length, width, height))
    dist = ((index_x - center[0]) ** 2 + (index_y - center[1]) ** 2 + (index_z - center[2]) ** 2) ** 0.5
    return 1 * (dist <= radius) - 1 * (dist <= radius - 1)


def testchange():
    arrayList = np.zeros((1, 10, 10))
    arrayList[0][1][1] = 1
    arrayList[0][0][1] = 2
    arrayList[0][1][0] = 3
    # init the list
    tupleList = []

    # rephrase ndarray to tuple
    for z in range(len(arrayList)):
        temp1 = []
        for y in range(len(arrayList[z])):
            temp2 = []
            for x in range(len(arrayList[z][y])):
                position = (x, y, z, arrayList[z][y][x])
                temp2.append(position)
            temp1.append(temp2)
        tupleList.append(temp1)

    a = []
    for x in tupleList:
        for y in x:
            a.extend(y)

    y, y_boundary = 0, 2
    x, x_boundary = 0, 3
    filmTuple_use = []
    for y_pos in range(y, y_boundary):
        for x_pos in range(x, x_boundary):
            filmTuple_use.append(tupleList[0][y_pos][x_pos])

    return tupleList


def testVisible():
    array = np.ones((10, 10))

    from matplotlib import pyplot as plt
    pos = np.where(array == 1)
    neu = np.where(array == 0)
    neg = np.where(array == -1)

    pos_x = pos[0]
    pos_y = pos[1]
    neu_x = neu[0]
    neu_y = neu[1]
    neg_x = neg[0]
    neg_y = neg[1]

    if len(array[0]) > 1000:
        img_length = len(array[0]) // 1000
        img_width = len(array) // 1000
    elif len(array[0]) > 100:
        img_length = len(array[0]) // 100
        img_width = len(array) // 100
    else:
        img_length = len(array[0])
        img_width = len(array)

    # set the size of point, 7.62 is 1 inch
    size = 7.62 * 7.62 * img_length

    fig, ax = plt.subplots(figsize=(img_length, img_width))
    # ax = fig.add_subplot(111)

    if len(pos_x) != 0:
        ax.scatter(pos_x, pos_y, s=size, c='blue', label='pos', marker=',')
    if len(neu_x) != 0:
        ax.scatter(neu_x, neu_y, s=size, c='green', label='neu', marker=',')
    if len(neg_x) != 0:
        ax.scatter(neg_x, neg_y, s=size, c='red', label='neg', marker=',')

    ax.legend(loc="upper right")
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.xaxis.set_ticks_position('top')
    ax.xaxis.set_label_position('top')

    plt.imshow(array, interpolation='nearest')

    plt.savefig("test.png")
    plt.show()


def testPic():
    from ExternalIO import visPlot
    data = np.zeros((100, 100))
    np.reshape(data, (-1,))
    data[2][2] = 1
    data[1][1] = -1

    picName = "test"

    visPlot(data, picName)


def testCutoff():
    from SimulatorFile.EnergyCalculator import interact2D

    film = np.zeros((1, 10, 10))
    film[0][2][2] = 1
    film[0][2][4] = 1
    film[0][1][2] = 1
    film[0][1][5] = 1
    film[0][0][2] = 1
    film[0][0][8] = 1
    film[0][1][7] = 1

    bacteria = np.zeros((1, 3, 3))
    bacteria[0][1][1] = 1
    bacteria[0][2][2] = -1

    result = interact2D("CUTOFF", 2, 2, film, bacteria, 1, 6)

    return result


def test_random():
    true_num = 0
    false_num = 0
    probability = 0.2

    for i in range(1000):
        result = _simple(probability)
        if result == 1:
            false_num += 1
        else:
            true_num += 1

    return "True is {}, False is {}".format(true_num, false_num)


def _simple(probability: float) -> bool:
    # check does probability set
    if probability is None:
        raise RuntimeError("Probability is not given")

    stick = np.random.choice([1, 0], 1, p=[probability, 1 - probability])

    return stick


def test_mp1():
    ncpus = int(os.environ.get('SLURM_CPUS_PER_TASK', default=1))
    pool = mp.Pool(processes=ncpus)
    data = [(11, 1), (22, 3), (41, 5, 6)]
    result = {}

    _cube = partial(cube, result=result)

    cubes = pool.map(_cube, data)
    print(cubes)
    cubes.sort()
    print(cubes)


def cube(x):
    return (x[0], x[0] ** 3)


def test_mp2():
    # init parameter for multiprocess
    ncpus = 10
    pool = mp.Pool(processes=ncpus)

    range_x = [i for i in range(50)]
    range_y = [i for i in range(50)]

    # prepare data for multiprocess, data is divided range into various parts, not exceed sqrt of ncpus can use
    part = int(np.floor(np.sqrt(ncpus)))
    data = []

    # double loop to prepare range x and range y
    range_x_list = [range_x[i:i + part] for i in range(0, len(range_x), part)]
    range_y_list = [range_y[i:i + part] for i in range(0, len(range_y), part)]

    # put combination into data
    for x in range_x_list:
        for y in range_y_list:
            data.append((x, y))

    result = pool.map(cube, data)

    # get the minimum result
    result.sort()
    result = result.pop(0)
    result, min_film = result[0], result[1]


def testCutoffFilm():
    film = np.arange(start=0, stop=10000)
    film = np.reshape(film, (100, 100))
    bacteria = np.ones((10, 10))

    startPoint = (10, 10)
    cutoff = 2

    r = _getCutoffFilm1D(film, startPoint, bacteria.shape, cutoff)
    print(r)


def _getCutoffFilm1D(film: ndarray, startPoint: Tuple[int, int], bacteriaSize: Tuple[int, int], cutoff: int) \
        -> List[ndarray]:
    """
    This function based on the start point and cut off value, general a list of film need to calculate energy
    """
    # init a empty list
    film_list = []

    # get the size of film
    if len(film.shape) == 2:
        filmSize = film.shape
    else:
        filmSize = film.shape[1:]

    # get the range on x and y axis
    x_start = max(0, startPoint[0] - cutoff)
    x_end = min(filmSize[1] - 1 - bacteriaSize[1], startPoint[0] + cutoff)
    y_start = max(0, startPoint[1] - cutoff)
    y_end = min(filmSize[0] - 1 - bacteriaSize[0], startPoint[1] + cutoff)

    # append corresponding film into list in 1D
    for x_s in range(x_start, x_end + 1):
        for y_s in range(y_start, y_end + 1):
            y = y_s + bacteriaSize[1]
            x = x_s + bacteriaSize[0]
            fy = film[y_s: y, x_s: x]
            film_list.append(fy)

    return film_list


def testSave():
    arrayList = np.zeros((1, 10, 10))
    arrayList[0][1][1] = 1
    arrayList[0][0][1] = 2
    arrayList[0][1][0] = 3

    a = 0.1
    b = True
    c = "Circle"

    info = [a, b, c, arrayList]

    # create file name
    file_name = ""
    for i in info[:-1]:
        file_name += str(i)
        file_name += "_"

    saveSurface(info, file_name[:-1])

    a = importSurface("{}.npy".format(file_name[:-1]))
    print(a)


    # name = 0.1

    # np.save("{}".format(name), np.array([name, arrayList], dtype=object))
    #
    # a = np.load("{}.npy".format(name), allow_pickle=True)


if __name__ == '__main__':
    # test_diamond()
    # test_random_choice()

    # a = A()
    # print(a.__dict__)
    # a.check()

    # a = 12
    # t()

    # test_excel()

    # _output()

    # test_simulation()

    # p = p()
    # t()

    # print(test3D())

    # print(testchange())

    # testVisible()

    # testPic()

    # print(testCutoff())

    # print(test_random())

    # test_mp2()

    # testCutoffFilm()

    testSave()
