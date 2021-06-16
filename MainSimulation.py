"""
This is the simulation demo program, take in the argument user put in and call appropriate function to generate
appropriate parameter and run the simulation and output the result into file
"""
from datetime import datetime
from typing import Tuple

import numpy as np

from Film import Film
from Bacteria import Bacteria
from ExternalIO import showMessage, writeLog
from FilmManager import FilmManager
from BacteriaManager import BacteriaManager
from openpyxl import Workbook
from openpyxl.utils import get_column_letter  # allows access to letters of each column


class Simulation:
    """
    This class is used for simulation
    """

    def __init__(self, simulationType: int, trail: int, dimension: int,
                 filmSeed: int, filmSurfaceSize: Tuple[int, int], filmSurfaceShape: str, filmSurfaceCharge: int,
                 filmDomainSize: Tuple[int, int], filmDomainShape: str, filmDomainConcentration: float,
                 bacteriaSeed: int, bacteriaSize: Tuple[int, int], bacteriaSurfaceShape: str,
                 bacteriaSurfaceCharge: int,
                 bacteriaDomainSize: Tuple[int, int], bacteriaDomainShape: str, bacteriaDomainConcentration: float,
                 filmNum: int, bacteriaNum: int, interval_x: int, interval_y: int):
        """
        Init the simulation class based on the input info
        Description of input info are shown in the HelpFile.txt
        """
        showMessage("start to init the simulation generator")

        # set variables
        self.trail = trail
        self.dimension = dimension
        self.simulationType = simulationType
        self.interval_x = interval_x
        self.interval_y = interval_y

        # init some variable
        self.filmManager = FilmManager(trail, dimension, filmSeed, filmSurfaceSize, filmSurfaceShape, filmSurfaceCharge,
                                       filmDomainSize, filmDomainShape, filmDomainConcentration, filmNum)

        self.bacteriaManager = BacteriaManager(trail, dimension, bacteriaSeed, bacteriaSize, bacteriaSurfaceShape,
                                               bacteriaSurfaceCharge, bacteriaDomainSize, bacteriaDomainShape,
                                               bacteriaDomainConcentration, bacteriaNum)
        self.startTime = datetime.now()
        self.output = self._init_output()

        # generate corresponding variable
        self.filmManager.generateFilm()
        self.bacteriaManager.generateBacteria()

        # write to log
        writeLog(self.__dict__)

    def runSimulate(self):
        """
        Based on the simulation type, do the corresponding simulation
        """
        # record the number of simulation did
        currIter = 0

        # type 1 simulation
        # only one film and one bacteria
        if self.simulationType == 1:
            self._simulate(currIter, self.filmManager.film[0], self.bacteriaManager.bacteria[0])

        # type 2 simulation
        elif self.simulationType == 2:
            # One film, multiple different bacteria, every bacteria scan the surface once
            for i in range(self.bacteriaManager.bacteriaNum):
                showMessage("This is type 2 simulation with simulation #: {}".format({i}))

                # start simulation
                self._simulate(currIter, self.filmManager.film[0], self.bacteriaManager.bacteria[currIter])
                currIter += 1

        # type 3 simulation
        elif self.simulationType == 3:
            # multiple different film, one bacteria, bacteria scan every surface once
            for i in range(self.filmManager.filmNum):
                showMessage("This is type 3 simulation with simulation #: {}".format({i}))

                # start simulation
                self._simulate(currIter, self.filmManager.film[currIter], self.bacteriaManager.bacteria[0])
                currIter += 1

    def _simulate(self, currIter: int, film: Film, bacteria: Bacteria):
        """
        This is the simulation function in this program, call function do the simulation and output the result
        Prerequisite: surface already generated
        """
        showMessage("Start to run simulation")

        # call simulation based on the simulation type
        if self.dimension == 2:
            result = self._interact2D(self.interval_x, self.interval_y, film, bacteria)
        elif self.dimension == 3:
            raise NotImplementedError

        # set the output
        self._output(result, currIter)

    def _init_output(self):
        """
        Init the out put excel file
        copy from the old code
        """
        # creates excel file
        wb = Workbook()
        ws1 = wb.create_sheet("Results", 0)

        # naming the columns in the worksheet
        ws1.cell(1, 1, "Surface Characteristics")
        ws1.cell(1, 2, "Bacteria Characteristics")
        ws1.cell(1, 3, "Film Seed # ")
        ws1.cell(1, 4, "Bacteria Seed # ")
        ws1.cell(1, 5, "Min Energy")
        ws1.cell(1, 6, "Min X")
        ws1.cell(1, 7, "Min Y")
        ws1.cell(1, 8, "Surface Charge at Min Energy")
        ws1.cell(1, 9, "Min Energy Gradient Strip")
        ws1.cell(1, 10, "Time used")
        ws1.cell(1, 11, "Histogram")

        # create numbering for histogram plot
        count = 0
        for i in range(12, 31):
            ws1.cell(1, i, count)
            ws1.cell(2, i, 0)
            count += 1

        # adjust column width to text length
        for i in range(ws1.max_column):
            text = ws1.cell(1, i + 1).value
            if type(text) == int:
                text = str(text)
                column_width = len(text) + 1
            else:
                column_width = len(text)
            ws1.column_dimensions[get_column_letter(i + 1)].width = column_width

        return (wb, ws1)

    def _output(self, result: Tuple, currIter: int):
        """
        Output the simulation result into a file
        Copy from old code with minor change
        """
        showMessage("Start to out put ......")

        # calculate the time use
        time_consume = (datetime.now() - self.startTime)
        time_consume = time_consume.seconds

        # rename the out put
        wb = self.output[0]
        ws1 = self.output[1]

        # split the result
        min_energy, min_x, min_y, min_energy_charge, min_charge, min_charge_x, min_charge_y = result

        # shows the gradient lines position
        grad_strip = min_x // 500

        # set the row position, 1 for the title of the whole sheet
        row_pos = 1 + currIter

        # write the result
        ws1.cell(row_pos, 1, str(self.filmManager.filmDomainShape) + " : " + str(self.filmManager.filmDomainSize))
        ws1.cell(row_pos, 2, str(self.bacteriaManager.bacteriaDomainShape) + " : " + str(self.bacteriaManager.bacteriaDomainSize))
        ws1.cell(row_pos, 3, self.filmManager.film[currIter].seed)
        ws1.cell(row_pos, 4, self.bacteriaManager.bacteria[currIter].seed)
        ws1.cell(row_pos, 5, min_energy)
        ws1.cell(row_pos, 6, min_x)
        ws1.cell(row_pos, 7, min_x)
        ws1.cell(row_pos, 8, min_energy_charge)
        ws1.cell(row_pos, 9, grad_strip)
        ws1.cell(row_pos, 10, time_consume)

        # if this is not the last iterator, update the time and return this

        # special count for simulation type 2
        # count number of min_energy locations at each gradient strip
        if self.simulationType == 2:
            for row_num in range(self.bacteriaManager.bacteriaNum):
                row = self.trail + row_num
                val_id = ws1.cell(row, 8).value
                val = ws1.cell(2, 10 + int(val_id)).value
                ws1.cell(2, 10 + int(val_id), int(val) + 1)

        # save the excel file into folder result
        name = "Simulation type {} trail {}_{}.xlsx".format(str(self.simulationType), self.trail,
                                                            datetime.now().strftime("%H-%M-%S"))
        file_path = "Result/" + name
        wb.save(file_path)

        showMessage("Output done")

    def _interact2D(self, interval_x: int, interval_y: int, film: Film, bacteria: Bacteria):
        """
        Do the simulation, scan whole film surface with bacteria
        The energy calculate only between bacteria surface and the film surface directly under the bacteria
        This code is copy from the old code with minor name change
        """
        showMessage("Start to interact ......")
        # shape of the bacteria
        shape = bacteria.shape

        # set the range
        range_x = np.arange(0, shape[0], interval_x)
        range_y = np.arange(0, shape[1], interval_y)

        # init some variable
        # randomly, just not negative
        min_energy = 1000
        min_charge = 1000
        min_energy_charge = 1000
        min_charge_x = 0
        min_charge_y = 0
        min_x = -1
        min_y = -1

        # change the bacteria surface into 1D
        bacteria_1D = np.reshape(bacteria.surfaceWithDomain, (-1))

        # scan through the surface and make calculation
        for x in range_x:
            for y in range_y:
                # set the x boundary and y boundary
                x_boundary = bacteria.width + x
                y_boundary = bacteria.length + y

                # check if bacteria surface is exceed range of film surface
                if x_boundary > shape[0] or y_boundary > shape[1]:
                    # if exceed the surface, go to next iteration
                    continue

                # do the calculation

                # change the corresponding film surface into 1D
                film_use = film.surfaceWithDomain[x: x_boundary, y: y_boundary]
                film_1D = np.reshape(film_use, (-1))

                # calculate energy, uses electrostatic energy formula, assuming that r = 1
                # WARNING: r should be change based on the height difference between film and bacteria
                energy = np.dot(film_1D, bacteria_1D)

                # count and unique
                unique, counts = np.unique(film_use, return_counts=True)

                # check the calculation result and change corresponding value
                if len(unique) == 1:
                    if unique[0] == -1:
                        charge = -counts[0]
                    else:
                        charge = counts[0]
                elif unique[0] == -1:
                    charge = -counts[0] + counts[1]
                elif unique[0] == 1:
                    charge = counts[0] - counts[1]

                if charge < min_charge:
                    min_charge = charge
                    min_charge_x = x
                    min_charge_y = y

                # find minimum energy and location
                if energy < min_energy:
                    min_energy = energy
                    min_x = x
                    min_y = y
                    min_energy_charge = charge

        # save the result
        result = (min_energy, min_x, min_y, min_energy_charge, min_charge, min_charge_x, min_charge_y)

        showMessage("Interact done")
        writeLog(result)

        return result

    def _interact3D(self):
        raise NotImplementedError
