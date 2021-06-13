"""
This is the simulation demo program, take in the argument user put in and call appropriate function to generate
appropriate parameter and run the simulation and output the result into file
"""
from typing import Tuple

import numpy as np

import Bacteria
import Film
from Bacteria import Bacteria2D
from Domain import DomainGenerator
from Film import FilmSurface2D
from openpyxl import Workbook
from openpyxl.utils import get_column_letter  # allows access to letters of each column


class Simulation:
    """
    This class is used for simulation
    """

    def __init__(self, simulationType: int, trail: int, dimension: int,
                 filmSeed: int, filmSurfaceSize: Tuple[int, int], filmSurfaceShape: str, filmSurfaceCharge: int,
                 filmDomainSize: str, filmDomainShape: str, filmDomainConcentration: float,
                 bacteriaSeed: int, bacteriaSize: Tuple[int, int], bacteriaSurfaceShape: str, bacteriaSurfaceCharge: int,
                 bacteriaDomainSize: str, bacteriaDomainShape: str, bacteriaDomainConcentration: float):
        """
        Init the simulation class based on the input info
        Description of input info are shown in the HelpFile.txt
        """
        # set variables
        self.trail = trail
        self.dimension = dimension
        self.simulationType = simulationType

        # set film variable
        self.filmSurfaceSize = filmSurfaceSize
        self.filmSurfaceShape = filmSurfaceShape
        self.filmSurfaceCharge = filmSurfaceCharge
        self.filmDomainSize = filmDomainSize
        self.filmDomainShape = filmDomainShape
        self.filmDomainConcentration = filmDomainConcentration

        # set bacteria variable
        self.bacteriaSize = bacteriaSize
        self.bacteriaSurfaceShape = bacteriaSurfaceShape
        self.bacteriaSurfaceCharge = bacteriaSurfaceCharge
        self.bacteriaDomainSize = bacteriaDomainSize
        self.bacteriaDomainShape = bacteriaDomainShape
        self.bacteriaDomainConcentration = bacteriaDomainConcentration

        # generate domain generator
        self.filmDomainGenerator = DomainGenerator(self.trail, filmSeed)
        self.bacteriaDomainGenerator = DomainGenerator(self.trail, bacteriaSeed)

        # init some variable
        self.film = None
        self.bacteria = None

    def generateAllSurface(self):
        """
        This function call appropriate function to generate the Film and Bacteria
        """
        # generate simulation surface
        if self.dimension == 2:
            # generate corresponding film and bacteria
            self.generate2DFilm()
            self.generate2DBacteria()

        elif self.dimension == 3:
            # generate corresponding film and bacteria
            raise NotImplementedError

    def generate2DFilm(self):
        """
        Generate 2D film
        """
        # generate 2D Film Surface
        self.film = FilmSurface2D(self.trail, self.filmSurfaceShape, self.filmSurfaceSize, self.filmSurfaceCharge,
                                  self.filmDomainGenerator, self.filmDomainShape, self.filmDomainSize,
                                  self.filmDomainConcentration)

    def generate2DBacteria(self):
        """
        Generate 2D bacteria
        """
        # generate 2D bacteria
        self.bacteria = Bacteria2D(self.trail, self.bacteriaSurfaceShape, self.bacteriaSize,
                                   self.bacteriaSurfaceCharge, self.bacteriaDomainGenerator, self.bacteriaDomainShape,
                                   self.bacteriaDomainSize, self.bacteriaDomainConcentration)

    def generateNewSurface(self, trail: int, seed: int, surfaceName: str):
        """
        This function generate new Film or Bacteria with given seed
        """
        # generate new domain generator
        domainGenerator = DomainGenerator(trail, seed)

        # check the surface want to generate
        if surfaceName.upper() == "FILM":
            # set new generator
            self.filmDomainGenerator = domainGenerator

            if self.dimension == 2:
                self.generate2DFilm()

        elif surfaceName.upper() == "BACTERIA":
            # set new generator
            self.bacteriaDomainGenerator = domainGenerator

            if self.dimension == 2:
                self.generate2DBacteria()

    def simulate(self, interval_x: int, interval_y: int):
        """
        This is the simulation function in this program, call function do the simulation and output the result
        Prerequisite: surface already generated
        """
        # check does surface generated
        if self.bacteria is None or self.film is None:
            self.generateAllSurface()

        # call simulation based on the simulation type
        if self.dimension == 2:
            self._interact2D(interval_x, interval_y)

        # set the output

    def _output(self):
        """
        Output the simulation result into a file
        Just copy from the old code
        """
        dirname = r""  # find directory of this file
        wb = Workbook()  # creates excel file
        ws1 = wb.create_sheet("Results", 0)

        # naming the columns in the worksheet
        ws1.cell(1, 1, "Surface Characteristics:")
        ws1.cell(1, 2, "Bacteria Characteristics:")
        ws1.cell(1, 3, "Seed # ")
        ws1.cell(1, 4, "Min Energy:")
        ws1.cell(1, 5, "Min X:")
        ws1.cell(1, 6, "Min Y:")
        ws1.cell(1, 7, "Surface Charge at Min Energy:")
        ws1.cell(1, 8, "Min Energy Gradient Strip: ")
        ws1.cell(1, 9, "Histogram: ")
        # create numbering for histogram plot
        count = 0
        for i in range(10, 31):
            ws1.cell(1, i, count)
            ws1.cell(2, i, 0)
            count += 1

        # adjust column width to text length
        for i in range(ws1.max_column):
            text = ws1.cell(1, i + 1).value
            if type(text) == int:
                text = str(text)
                column_width = len(text) + 1
            else:
                column_width = len(text)
            ws1.column_dimensions[get_column_letter(i + 1)].width = column_width

    def _interact2D(self, interval_x: int, interval_y: int):
        """
        Do the simulation, scan whole film surface with bacteria
        The energy calculate only between bacteria surface and the film surface directly under the bacteria
        """
        # shape of the bacteria
        shape = self.bacteria.shape

        # set the range
        range_x = np.arange(0, shape[0], interval_x)
        range_y = np.arange(0, shape[1], interval_y)

        # init some variable
        # randomly, just not negative
        min_energy = 1000
        min_charge = 1000
        min_energy_charge = 1000
        min_charge_x = 0
        min_charge_y = 0
        min_x = -1
        min_y = -1

        # change the bacteria surface into 1D
        bacteria_1D = np.reshape(self.bacteria.surfaceWithDomain, (-1))

        # scan through the surface and make calculation
        for x in range_x:
            for y in range_y:
                # set the x boundary and y boundary
                x_boundary = self.bacteria.width + x
                y_boundary = self.bacteria.length + y

                # check if bacteria surface is exceed range of film surface
                if x_boundary > shape[0] or y_boundary > shape[1]:
                    # if exceed the surface, go to next iteration
                    continue

                # do the calculation

                # change the corresponding film surface into 1D
                film_use = self.film.surfaceWithDomain[x: x_boundary, y: y_boundary]
                film_1D = np.reshape(film_use, (-1))

                # calculate energy, uses electrostatic energy formula, assuming that r = 1
                # WARNING: r should be change based on the height difference between film and bacteria
                energy = np.dot(film_1D, bacteria_1D)

                # count and unique
                unique, counts = np.unique(film_use, return_counts=True)

                # check the calculation result and change corresponding value
                if len(unique) == 1:
                    if unique[0] == -1:
                        charge = -counts[0]
                    else:
                        charge = counts[0]
                elif unique[0] == -1:
                    charge = -counts[0] + counts[1]
                elif unique[0] == 1:
                    charge = counts[0] - counts[1]

                if charge < min_charge:
                    min_charge = charge
                    min_charge_x = x
                    min_charge_y = y

                # find minimum energy and location
                if energy < min_energy:
                    min_energy = energy
                    min_x = x
                    min_y = y
                    min_energy_charge = charge

        return (min_energy, min_x, min_y, min_energy_charge, min_charge, min_charge_x, min_charge_y)

    def _interact3D(self):
        raise NotImplementedError
