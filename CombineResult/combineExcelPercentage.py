"""
This file:
- Generates the graph of "Number of stuck bacteria at certain time VS Total number of bacteria"
- Gets several dynamic result files with same time step and combines them

* "Time" is an independent variable
* "Stuck bacteria number over time" is a dependent variable
"""

"""
Procedure for using this:
1. Put all dynamic results you want to combine in a folder called DynamicResult, 
2. Make sure this folder is under same root as this file,
3. Run this file to output your combined results in the folder DynamicResult.
"""

import os
from openpyxl import load_workbook, Workbook

folder_path = "../DynamicResult"

file_list = os.listdir(folder_path)
file_list.sort()

result = []
# generate title
title = []
bactNum = []

# init a new workbook
wb = Workbook()
ws1 = wb.create_sheet("Combine", 0)
timeStep = None

for file in file_list:
    if "trail" in file:
        filepath = "{}/{}".format(folder_path, file)
        workbook = load_workbook(filename=filepath)
        sheet = workbook["Results"]

        # get total time step
        if timeStep is None:
            timeStep = sheet["L"]
            result.append(timeStep)
            title.append("Time step")

        total_bact_col_name = "I"
        stuck_bact_col_name = "N"

        # get column for total bacteria and stuck bacteria
        total_bact_col = sheet[total_bact_col_name]
        stuck_bact_col = sheet[stuck_bact_col_name]

        # Record result
        bact_num = total_bact_col[2].value
        title.append("{} bact stuck %".format(bact_num))
        bactNum.append(bact_num)
        result.append(stuck_bact_col)


# write result into ws

# write title
ws1.append(title)

# generate rwo list, read every part in
for row_num in range(1, len(timeStep)):
    rowList = []
    n = -1
    for col in result:
        if n == -1:
            rowList.append(col[row_num].value)
        else:
            # if this is time step, just append
            rowList.append(col[row_num].value / bactNum[n])
        n += 1

    ws1.append(rowList)


# save result
name = "Dynamic_combine_percen.xlsx"

file_path = "{}/{}".format(folder_path, name)

wb.save(file_path)




