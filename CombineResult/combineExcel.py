"""
This file is used for generate Graph: number of stuck bacteria at certain time vs total bacteria number
Get several dynamic result files with same time step and convert it to
time is independent variable, dependent variable is stuck bacteria number over time
"""

"""
Procedure for using this:
1. Put all dynamic result want to combine in a folder called DynamicResult, make sure this folder under same root as this file
2. Run this file and you can get combined result in the folder DynamicResult.
"""

import os
from openpyxl import load_workbook, Workbook

folder_path = "../DynamicResult"

file_list = os.listdir(folder_path)
file_list.sort()

result = []

# init a new workbook
wb = Workbook()
ws1 = wb.create_sheet("Combine", 0)
timeStep = None

for file in file_list:
    if "trail" in file:
        filepath = "{}/{}".format(folder_path, file)
        workbook = load_workbook(filename=filepath)
        sheet = workbook["Results"]

        # get total time step
        if timeStep is None:
            timeStep = sheet["L"]
            result.append(timeStep)

        total_bact_col_name = "I"
        stuck_bact_col_name = "N"

        # get column for total bacteria and stuck bacteria
        total_bact_col = sheet[total_bact_col_name]
        stuck_bact_col = sheet[stuck_bact_col_name]

        # Record result
        result.append(total_bact_col)
        result.append(stuck_bact_col)


# write result into ws
# generate rwo list, read every part in
for row_num in range(len(timeStep)):
    rowList = []
    for col in result:
        rowList.append(col[row_num].value)
    ws1.append(rowList)


# save result
name = "Dynamic_combine.xlsx"

file_path = "{}/{}".format(folder_path, name)

wb.save(file_path)




