"""
This is energy scan simulator
"""
from datetime import datetime
from typing import Tuple, Union, List, Dict

import numpy as np
from numpy import ndarray
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from SimulatorFile.EnergyCalculator import interact2D, interact3D
from ExternalIO import showMessage, writeLog, saveResult
from openpyxl import Workbook
from openpyxl.utils import get_column_letter  # allows access to letters of each column
from SimulatorFile.Simulator import Simulator


class EnergySimulator(Simulator):
    """
    This class is used for bacteria scan film surface energy simulation
    """
    interactType: Union[None, str]

    def __init__(self, trail: int, dimension: int,
                 filmSeed: int, filmSurfaceSize: Union[Tuple[int, int], Tuple[int, int, int]], filmSurfaceShape: str,
                 filmSurfaceCharge: int, filmDomainSize: Tuple[int, int], filmDomainShape: str,
                 filmDomainConcentration: float, filmDomainChargeConcentration: float,
                 bacteriaSeed: int, bacteriaSize: Union[Tuple[int, int], Tuple[int, int, int]], bacteriaSurfaceShape: str,
                 bacteriaSurfaceCharge: int,
                 bacteriaDomainSize: Tuple[int, int], bacteriaDomainShape: str, bacteriaDomainConcentration: float,
                 bacteriaDomainChargeConcentration: float,
                 filmNum: int, bacteriaNum: int, intervalX: int, intervalY: int, parameters: Dict) -> None:
        """
        Init the simulation class based on the input info
        Description of input info are shown in the HelpFile.txt
        """
        simulatorType = 1

        # get simulationType
        if "simulationType" not in parameters:
            raise RuntimeError("simulation type is not enetered")
        else:
            simulationType = parameters["simulationType"]

        # if simulation type is 1, fix the bacteria number to 1
        if simulationType == 1:
            bacteriaNum = 1

        # set some variable
        self.interactType = None
        self.cutoff = -1

        # call parent to generate simulator
        Simulator.__init__(self, simulationType, trail, dimension, simulatorType,
                           filmSeed, filmSurfaceSize, filmSurfaceShape, filmSurfaceCharge,
                           filmDomainSize, filmDomainShape, filmDomainConcentration, filmDomainChargeConcentration,
                           bacteriaSeed, bacteriaSize, bacteriaSurfaceShape, bacteriaSurfaceCharge,
                           bacteriaDomainSize, bacteriaDomainShape, bacteriaDomainConcentration,
                           bacteriaDomainChargeConcentration,
                           filmNum, bacteriaNum, intervalX, intervalY, parameters)

    def runSimulate(self) -> None:
        """
        Based on the simulation type, do the corresponding simulation
        """
        writeLog("This is runSimulation in Simulation")
        showMessage("Start to run simulation baed on simulation type")
        writeLog(self.__dict__)

        # record the number of simulation did
        currIter = 0

        # init the end of iterator
        end = False

        # type 1 simulation
        # only one film and one bacteria
        if self.simulationType == 1:
            end = True
            self._simulate(currIter, self.filmManager.film[0].surfaceWithDomain,
                           self.bacteriaManager.bacteria[0].surfaceWithDomain, end)

        # type 2 simulation
        elif self.simulationType == 2:
            # One film, multiple different bacteria, every bacteria scan the surface once
            for i in range(self.bacteriaManager.bacteriaNum):
                showMessage("This is type 2 simulation with simulation #: {}".format(i))

                # change end indicator
                if i == self.bacteriaManager.bacteriaNum - 1:
                    end = True

                # start simulation
                self._simulate(currIter, self.filmManager.film[0].surfaceWithDomain,
                               self.bacteriaManager.bacteria[currIter].surfaceWithDomain, end)
                currIter += 1

        # type 3 simulation
        elif self.simulationType == 3:
            # multiple different film, one bacteria, bacteria scan every surface once
            for i in range(self.filmManager.filmNum):
                showMessage("This is type 3 simulation with simulation #: {}".format({i}))

                # change end indicator
                if i == self.filmManager.filmNum - 1:
                    end = True

                # start simulation
                self._simulate(currIter, self.filmManager.film[currIter].surfaceWithDomain,
                               self.bacteriaManager.bacteria[0].surfaceWithDomain, end)
                currIter += 1
        else:
            raise RuntimeError("Wrong simulation type")

    def _simulate(self, currIter: int, film: ndarray, bacteria: ndarray, end: bool) -> None:
        """
        This is the simulation function in this program, call function do the simulation and output the result
        Prerequisite: surface already generated
        """
        writeLog("This is _simulate in Simulation")
        showMessage("Start to run simulation")
        writeLog("self is: {}, currIter is: {}, film is: {}, bacteria is: {}, end is: {}".format(
            self.__dict__, currIter, film, bacteria, end))

        # check does cutoff value set
        if self.interactType.upper() in ["CUTOFF", "CUT-OFF"]:
            if self.cutoff < 0:
                raise RuntimeError("Cutoff value is not assign or not assign properly")
            else:
                cutoff = self.cutoff
        else:
            cutoff = 0

        # call simulation based on the simulation type
        if self.dimension == 2:
            result = interact2D(self.interactType, self.intervalX, self.intervalY, film, bacteria, currIter, cutoff)
        elif self.dimension == 3:
            result = interact3D(self.interactType, self.intervalX, self.intervalY, film, bacteria, currIter, cutoff)
        else:
            raise RuntimeError("Wrong dimension in _simulate")

        showMessage("Interact done")

        # set the output
        self._output(result, currIter, end)

    def _initOutput(self) -> Tuple[Workbook, Union[WriteOnlyWorksheet, Worksheet]]:
        """
        Init the out put excel file
        copy from the old code with separate seed into film seed and bacteria seed
        """
        writeLog("This is _initOutput in Simulation")
        showMessage("Init the output")
        writeLog(self.__dict__)

        # creates excel file
        wb = Workbook()
        ws1 = wb.create_sheet("Results", 0)

        # naming the columns in the worksheet
        ws1.cell(1, 1, "Trail")
        ws1.cell(1, 2, "Dimension")
        ws1.cell(1, 3, "Film shape and size")
        ws1.cell(1, 4, "Film domain shape and size")
        ws1.cell(1, 5, "Bacteria shape and size")
        ws1.cell(1, 6, "Bacteria domain shape and size")
        ws1.cell(1, 7, "Film Seed # ")
        ws1.cell(1, 8, "Film real domain concentration")
        ws1.cell(1, 9, "Bacteria Seed # ")
        ws1.cell(1, 10, "Bacteria real domain concentration")
        ws1.cell(1, 11, "Min Energy")
        ws1.cell(1, 12, "Min X")
        ws1.cell(1, 13, "Min Y")
        ws1.cell(1, 14, "Surface Charge at Min Energy")
        ws1.cell(1, 15, "Min Energy Gradient Strip")
        ws1.cell(1, 16, "Time used (s)")
        ws1.cell(1, 17, "Interact type")

        # if simulation type is 2, do the count
        if self.simulationType == 2:
            # create numbering for histogram plot
            count = 0
            # number is how many strip
            number = 20
            for i in range(18, 18 + number):
                ws1.cell(1, i, count)
                ws1.cell(2, i, 0)
                count += 1

        # adjust column width to text length
        for i in range(ws1.max_column):
            text = ws1.cell(1, i + 1).value
            if type(text) == int:
                text = str(text)
                column_width = len(text) + 1
            else:
                column_width = len(text)
            ws1.column_dimensions[get_column_letter(i + 1)].width = column_width

        return (wb, ws1)

    def _output(self, result: Tuple, currIter: int, end: bool) -> None:
        """
        Output the simulation result into a file
        Copy from old code with minor change
        """
        writeLog("This is _output in Simulation")
        showMessage("Start to write result into out put")
        writeLog("self is: {}, result is: {}, currIter is: {}, end is: {}".format(
            self.__dict__, result, currIter, end))

        # calculate the time use
        time_consume = (datetime.now() - self.startTime)
        time_consume = time_consume.seconds

        # rename the out put
        wb = self.output[0]
        ws1 = self.output[1]

        # split the result
        min_energy, min_x, min_y, min_energy_charge, min_charge, min_charge_x, min_charge_y = result

        # shows the gradient lines position
        grad_strip = min_x // 500

        # set the row position, 2 for the title of the whole sheet
        row_pos = 2 + currIter

        # write the result
        ws1.cell(row_pos, 1, str(self.trail))
        ws1.cell(row_pos, 2, self.dimension)
        ws1.cell(row_pos, 3, str(self.filmManager.filmSurfaceShape) + " : " + str(self.filmManager.filmSurfaceSize))
        ws1.cell(row_pos, 4, str(self.filmManager.filmDomainShape) + " : " + str(self.filmManager.filmDomainSize))
        ws1.cell(row_pos, 5,
                 str(self.bacteriaManager.bacteriaSurfaceShape) + " : " + str(self.bacteriaManager.bacteriaSize))
        ws1.cell(row_pos, 6,
                 str(self.bacteriaManager.bacteriaDomainShape) + " : " + str(self.bacteriaManager.bacteriaDomainSize))

        if self.simulationType == 3:
            ws1.cell(row_pos, 7, self.filmManager.film[currIter].seed)
            ws1.cell(row_pos, 8, self.filmManager.film[currIter].realDomainConc)
            ws1.cell(row_pos, 9, self.bacteriaManager.bacteria[0].seed)
            ws1.cell(row_pos, 10, self.bacteriaManager.bacteria[0].realDomainConc)

        else:
            ws1.cell(row_pos, 7, self.filmManager.film[0].seed)
            ws1.cell(row_pos, 8, self.filmManager.film[0].realDomainConc)
            ws1.cell(row_pos, 9, self.bacteriaManager.bacteria[currIter].seed)
            ws1.cell(row_pos, 10, self.bacteriaManager.bacteria[currIter].realDomainConc)

        ws1.cell(row_pos, 11, min_energy)
        ws1.cell(row_pos, 12, min_x)
        ws1.cell(row_pos, 13, min_x)
        ws1.cell(row_pos, 14, min_energy_charge)
        ws1.cell(row_pos, 15, grad_strip)
        ws1.cell(row_pos, 16, time_consume)
        if self.interactType.upper() == 'DOT':
            ws1.cell(row_pos, 17, self.interactType)
        else:
            ws1.cell(row_pos, 17, "{}: {}".format(self.interactType, self.cutoff))

        # if this is not the last iterator, update the time and return this
        if not end:
            self.startTime = datetime.now()
            return None

        # generate time and record it
        date = datetime.now().strftime("%m_%d")
        time = datetime.now().strftime("%H-%M-%S")

        # save no count first
        # save the excel file into folder result
        name = "EnergyScan_Type_{}_trail_{}-{}-{}_no_count.xlsx".format(str(self.simulationType), self.trail, date, time)
        file_path = "Result/" + name

        # call function in ExternalIO to save workbook
        saveResult(wb, file_path)

        # special count for simulation type 2
        # count number of min_energy locations at each gradient strip
        if self.simulationType == 2:
            showMessage("WARNING: Potential bug here")
            for row_num in range(self.bacteriaManager.bacteriaNum):
                row = 2 + row_num
                val_id = ws1.cell(row, 15).value

                # check the value read from column 11
                if val_id < 0:
                    continue
                val = ws1.cell(2, 18 + int(val_id)).value
                ws1.cell(2, 18 + int(val_id), int(val) + 1)

        # save the excel file into folder result
        name = "EnergyScan_Type_{}_trail_{}-{}-{}_count.xlsx".format(str(self.simulationType), self.trail, date, time)
        file_path = "Result/" + name

        # call function in ExternalIO to save workbook
        saveResult(wb, file_path)

