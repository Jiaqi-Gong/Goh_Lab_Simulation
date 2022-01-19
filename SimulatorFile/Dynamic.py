"""
This program:
- Simulates 3D Dynamic bacteria movement
"""
from datetime import datetime
from typing import Tuple, Union, List, Dict

from numpy import ndarray
from openpyxl import Workbook
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
import numpy as np
import scipy.optimize
import pandas as pd

from ExternalIO import writeLog, showMessage, saveResult, timstepPlot, monoExp, timeMonitor
from SimulatorFile.Simulator import Simulator
from BacteriaFile.BacteriaMovement import BacteriaMovementGenerator


class DynamicSimulator(Simulator):

    def __init__(self, trail: int, dimension: int,
                 filmSeed: int, filmSurfaceSize: Tuple[int, int], filmSurfaceShape: str, filmSurfaceCharge: int,
                 filmDomainSize: Tuple[int, int], filmDomainShape: str, filmDomainConcentration: float,
                 filmDomainChargeConcentration: float,
                 bacteriaSeed: int, bacteriaSize: Tuple[int, int], bacteriaSurfaceShape: str,
                 bacteriaSurfaceCharge: int,
                 bacteriaDomainSize: Tuple[int, int], bacteriaDomainShape: str, bacteriaDomainConcentration: float,
                 bacteriaDomainChargeConcentration: float,
                 filmNum: int, bacteriaNum: int, intervalX: int, intervalY: int, filmNeutralDomain: bool,
                 bacteriaNeutralDomain: bool, parameters: Dict, preparedSurface: ndarray = None) -> None:
        """
        Init the simulation class based on the input info
        Description of input info are shown in the HelpFile.txt
        """
        # check film and bacteria type
        if dimension != 3:
            raise RuntimeError("The film and bacteria dimension should be 3D for dynamic simulation.")

        # set some variable
        self.probabilityType = None
        self.timeStep = None
        self.bacteriaMovementSeed = None
        self.dumpStep = None
        self.unstuck = None
        self.generateDomain = parameters["generateDomain"]

        if parameters["unstuck"]:
            self.unstuckProbability = None

        # simulation type is not applicable for dynamic simulator for now, maybe in the future this can be use to do sth
        simulationType = -1

        # based on type, set parameter
        if parameters["probabilityType"].upper() == "SIMPLE":
            self.probability = None
        elif parameters["probabilityType"].upper() == "POISSON":
            self.Lambda = None
        elif parameters["probabilityType"].upper() == "BOLTZMANN":
            self.temperature = None
            self.energy = None

        # call parent to generate simulator
        simulatorType = 2
        Simulator.__init__(self, simulationType, trail, dimension, simulatorType,
                           filmSeed, filmSurfaceSize, filmSurfaceShape, filmSurfaceCharge,
                           filmDomainSize, filmDomainShape, filmDomainConcentration, filmDomainChargeConcentration,
                           bacteriaSeed, bacteriaSize, bacteriaSurfaceShape, bacteriaSurfaceCharge,
                           bacteriaDomainSize, bacteriaDomainShape, bacteriaDomainConcentration,
                           bacteriaDomainChargeConcentration,
                           filmNum, bacteriaNum, intervalX, intervalY, filmNeutralDomain, bacteriaNeutralDomain,
                           parameters, preparedSurface, self.generateDomain)

    def runSimulate(self) -> None:
        """
        This function do the simulation
        Implement in the super class abstract method
        """

        # raise NotImplementedError

        # set some variable, will replaced by user input
        z_restriction = 8
        bacteriaShape = self.bacteriaManager.bacteriaSurfaceShape

        # create bacteria movement generator
        # film and bacteria should be 3D and all film and bacteria should be same, only diff is domain distribution
        film = self.filmManager.film[0]
        bacteria = self.bacteriaManager.bacteria[0]

        # get the parameter need for movement generator
        filmSize = (film.length, film.width, film.height)
        bacteriaSize = (bacteria.length, bacteria.width, bacteria.height)
        bactMoveGenerator = BacteriaMovementGenerator(z_restriction, self.bacteriaMovementSeed, bacteriaShape,
                                                      filmSize, bacteriaSize)

        # init position for every bacteria
        for bacteria in self.bacteriaManager.bacteria:
            bacteria.position = bactMoveGenerator.initPosition()

        # record the first time
        result = [len(self.bacteriaManager.freeBacteria), len(self.bacteriaManager.stuckBacteria)]
        self._output(result, 0, False)

        showMessage("Beginning Dynamic simulator...")

        # do the simulation
        end = False
        for currIter in range(1, self.timeStep):
            # check the end of timeStep
            if currIter == self.timeStep - 1:
                end = True

            # do the simulation
            self._simulate(bactMoveGenerator)
            result = [len(self.bacteriaManager.freeBacteria), len(self.bacteriaManager.stuckBacteria)]

            # update the output based on the dump step
            if currIter % self.dumpStep == 0:
                self._output(result, currIter, end)

        # save the last simulate, if it is not saved in the loop
        if (self.timeStep - 1) % self.dumpStep != 0:
            self._output(result, self.timeStep - 1, end)

        showMessage("Dynamic Simulation: Complete.")

    def _initOutput(self) -> Tuple[Workbook, Union[WriteOnlyWorksheet, Worksheet]]:
        """
        This function init the format and content need to out put
        Implement in the super class abstract method
        """
        writeLog("This is _initOutput in Dynamic simulator.")
        showMessage("Initializing Dynamic Simulation output...")
        writeLog(self.__dict__)

        # creates excel file
        wb = Workbook()
        ws1 = wb.create_sheet("Results", 0)

        # naming the columns in the worksheet
        ws1.cell(1, 1, "Trail")
        ws1.cell(1, 2, "Film shape and size")
        ws1.cell(1, 3, "Film domain shape and size")
        ws1.cell(1, 4, "Film Seed # ")
        ws1.cell(1, 5, "Bacteria Movement generator seed")
        ws1.cell(1, 6, "Bacteria shape and size")
        ws1.cell(1, 7, "Bacteria domain shape and size")
        ws1.cell(1, 8, "Start Bacteria Seed # ")
        ws1.cell(1, 9, "Total bacteria number")
        ws1.cell(1, 10, "Time used (s)")
        ws1.cell(1, 11, "Probability type")
        ws1.cell(1, 12, "Time step")
        ws1.cell(1, 13, "Free bacteria number")
        ws1.cell(1, 14, "Stuck bacteria number")
        last_col = 14

        if self.probabilityType.upper() == "SIMPLE":
            ws1.cell(1, 15, "Probability")
            last_col += 1
        elif self.probabilityType.upper() == "POISSON":
            ws1.cell(1, 15, "Lambda value")
            last_col += 1
        elif self.probabilityType.upper() == "BOLTZMANN":
            ws1.cell(1, 15, "Temperature")
            ws1.cell(1, 16, "Energy")
            last_col += 2

        ws1.cell(1, last_col + 1, "Unstuck")
        last_col += 1
        # if it can unstuck
        if self.unstuck:
            ws1.cell(1, last_col + 1, "Stuck probability")
            last_col += 1

        # calculate the number of stuck / number of total
        ws1.cell(1, last_col + 1, "Number of stuck / number of total")

        return (wb, ws1)

    def _output(self, result: List[int], currIter: int, end: bool) -> None:
        """
        This function generate the info need to output
        Implement in the super class abstract method
        """
        writeLog("This is _output in Dynamic simulator.")
        writeLog("self is: {}, result is: {}, currIter is: {}, end is: {}".format(
            self.__dict__, result, currIter, end))

        # calculate the time use
        time_consume = (datetime.now() - self.startTime)
        time_consume = time_consume.seconds

        # rename the out put
        wb = self.output[0]
        ws1 = self.output[1]

        # set the row position, 2 for the title of the whole sheet
        row_pos = 2 + currIter

        # get the result
        freeBactNum, stuckBactNum = result

        # write the result
        ws1.cell(row_pos, 1, self.trail)
        ws1.cell(row_pos, 2, str(self.filmManager.filmSurfaceShape) + " : " + str(self.filmManager.filmSurfaceSize))
        ws1.cell(row_pos, 3, str(self.filmManager.filmDomainShape) + " : " + str(self.filmManager.filmDomainSize))
        ws1.cell(row_pos, 4, self.filmManager.film[0].seed)
        ws1.cell(row_pos, 5, self.bacteriaMovementSeed)
        ws1.cell(row_pos, 6,
                 str(self.bacteriaManager.bacteriaSurfaceShape) + " : " + str(self.bacteriaManager.bacteriaSize))
        ws1.cell(row_pos, 7,
                 str(self.bacteriaManager.bacteriaDomainShape) + " : " + str(self.bacteriaManager.bacteriaDomainSize))
        ws1.cell(row_pos, 8, self.bacteriaManager.bacteriaSeed)
        ws1.cell(row_pos, 9, self.bacteriaNum)
        ws1.cell(row_pos, 10, time_consume)
        ws1.cell(row_pos, 11, self.probabilityType)
        ws1.cell(row_pos, 12, currIter)
        ws1.cell(row_pos, 13, freeBactNum)
        ws1.cell(row_pos, 14, stuckBactNum)
        last_col = 14

        if self.probabilityType.upper() == "SIMPLE":
            ws1.cell(row_pos, 15, self.probability)
            last_col += 1
        elif self.probabilityType.upper() == "POISSON":
            ws1.cell(row_pos, 15, self.Lambda)
            last_col += 1
        elif self.probabilityType.upper() == "BOLTZMANN":
            ws1.cell(row_pos, 15, self.temperature)
            ws1.cell(row_pos, 16, self.energy)
            last_col += 2

        ws1.cell(row_pos, last_col + 1, self.unstuck)
        last_col += 1

        # if it can unstuck
        if self.unstuck:
            ws1.cell(row_pos, last_col + 1, self.unstuckProbability)
            last_col += 1

        # calculate the number of stuck / number of total
        ws1.cell(row_pos, last_col + 1, len(self.bacteriaManager.stuckBacteria) / self.bacteriaNum)

        # if this is not the last iterator, update the time and return this
        if not end:
            self.startTime = datetime.now()
            return None

        # save the excel file into folder result
        date = {"day": datetime.now().strftime("%m_%d"),
                "current_time": datetime.now().strftime("%H-%M-%S")}

        name = "Dynamic_trail_{}-{}-{}.xlsx".format(self.trail,
                                                    date["day"],
                                                    date["current_time"])
        file_path = "Result/ResultDynamic/" + name

        # call function in ExternalIO to save workbook
        saveResult(wb, file_path)

        # read the excel file to generate image
        # read the first sheet
        master_sheet = pd.read_excel(file_path, sheet_name=0, index_col=None)

        # determine the column of stuck bacteria and timestep
        timestep = master_sheet["Time step"]
        stuck_bacteria = master_sheet["Stuck bacteria number"]

        # now calculate the equilibrium amount of bacteria stuck in the film, since this is the final step
        equilibrium, param = self._calcEquilibrium(timestep, stuck_bacteria)

        # we only need to record the equilibrium amount of bacteria if the unstuck is true
        if self.unstuck:
            ws1.cell(1, 20, "equilibrium bacteria stuck")
            ws1.cell(2, 20, equilibrium)
            # call function in ExternalIO to save workbook
            saveResult(wb, file_path)

        # generate a graph
        timstepPlot(timestep, stuck_bacteria, param, date)

    @timeMonitor
    def _simulate(self, bactMoveGenerator: BacteriaMovementGenerator) -> None:
        """
        This function do the simulate
        """
        """
        1. Call initposition function in bact movement give all bacteria an init position
        2. Do a loop, in each loop, call stickOrNot and give new position unti the end of timeStep
        3. If bacteria is stick, change it from free list to stick list
        4. At the end, give the length of free and length of stick
        """
        # if no free bacteria, do nothing
        if not self.unstuck:
            if len(self.bacteriaManager.freeBacteria) == 0:
                return None

        # if probability type is Boltzmann, need to update self.temperature and self.energy
        # will be implement in the future

        # init two list for temp record
        free_bact = self.bacteriaManager.freeBacteria[:]
        stuck_bact = self.bacteriaManager.stuckBacteria[:]

        self._stickBact(bactMoveGenerator, free_bact, stuck_bact)

        # if can unstack, call function to unstuck bacteria
        if self.unstuck:
            self._unstuckBact(bactMoveGenerator, free_bact, stuck_bact)

        # updated free and stuck bacteria to manager
        self.bacteriaManager.freeBacteria = free_bact
        self.bacteriaManager.stuckBacteria = stuck_bact

    @timeMonitor
    def _stickBact(self, bactMoveGenerator, free_bact, stuck_bact):
        """
        A helper function to decide every bacteria stick or not
        """
        # loop all free bacteria
        for bact in self.bacteriaManager.freeBacteria:
            # get next position based on probability type
            if self.probabilityType.upper() == "SIMPLE":
                bactNextPos = bactMoveGenerator.nextPosition(self.probabilityType, bact.position,
                                                             probability=self.probability)
            else:
                raise RuntimeError(
                    "This is _interact in Dynamic simulator, the input probability type is not implement")

            # based on next position, move the bacteria
            if bactNextPos is False:
                # bacteria is stuck, move from free list to stuck list
                free_bact.remove(bact)
                stuck_bact.append(bact)
            else:
                # bacteria is not stuck, update the position
                bact.position = bactNextPos

    @timeMonitor
    def _unstuckBact(self, bactMoveGenerator, free_bact, stuck_bact):
        """
        A helper function for unstuck bacteria
        Loop all stuck bacteria and decide free or not
        """
        for bact in self.bacteriaManager.stuckBacteria:
            # get to see if bacteria can free
            if self.probabilityType.upper() == "SIMPLE":
                bactStatus = bactMoveGenerator.unstuckBacteria(self.probabilityType, self.unstuckProbability)
            else:
                raise RuntimeError(
                    "This is _interact in Dynamic simulator, the input probability type is not implement")

            # based on bact status, move bacteria in list
            if bactStatus:
                bactMoveGenerator.reliefOccupy(bact.position)
                stuck_bact.remove(bact)
                free_bact.append(bact)

    def _calcEquilibrium(self, timestep: List, stuck_bacteria: List) -> List:
        """
        This function calculates the equilibrium bacteria amount
        """
        p0 = (2000, .1, 50)  # start with values near those we expect
        params, cv = scipy.optimize.curve_fit(monoExp, timestep, stuck_bacteria, p0)
        m, t, b = params

        # y = -me^(-tx) + b
        # the equilibrium number would be the b value
        equilibrium = b

        return [equilibrium, params]
